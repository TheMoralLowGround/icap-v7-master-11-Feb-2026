"""
Organization: AIDocbuilder Inc.
File: scripts/RAJson.py
Version: 6.0

Authors:
    - Vivek - Initial implementation
    - Nayem - Code optimization

Last Updated By: Nayem
Last Updated At: 2024-12-04

Description:
    This script parse and process XML files to generate ra_json by extracting elements and
    mapping each element type to output name in the JSON.

Dependencies:
    - load_workbook from openpyxl
    - get_column_letter from openpyxl.utils
    - EmailParsedDocument, TrainParsedDocument from core.models

Main Features:
    - Extracts document, page, and element data recursively from XML layout files.
    - Assigns IDs to nested elements for structured identification.
    - Handle different document types including PDFs, Word documents, and Excel files.
    - Organizes Excel sheet data into a structured dictionary by columns and rows.
    - Retrieves document metadata for customized processing.
    - Processes nested elements like lines, words, and cells.
    - Generates a nested dictionary for further processing or JSON storage.
"""

import datetime
import json
import os
import xml.etree.ElementTree as ET
from copy import deepcopy
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from core.models import EmailParsedDocument, TrainParsedDocument, ApplicationSettings
from dashboard.models import Profile


def json_default(value):
    """Default value function for JSON dump"""
    if isinstance(value, (datetime.time, datetime.date, datetime.datetime)):
        return value.isoformat()
    raise ValueError(f"{type(value)} is not JSON Serializable.")


class RAJson:
    """
    NODES is dictionary where every key is Type of element, we need to extract from XML file. Value associated to each key is the name to be outputted in resulting JSON File.
    For example: Script will look for 'L' elements and put into json with type 'line'.
    """

    NODES = {
        "Block": "block",
        "Table": "table",
        "Row": "row",
        "Cell": "cell",
        "Para": "paragraph",
        "Title": "paragraph",
        "L": "line",
        "W": "word",
    }

    NODE_KEYS = list(NODES.keys())

    def __init__(
        self,
        batch_id,
        input_path,
        info_holder_data,
        valid_page_types,
        excel_read_only=True,
        read_checkbox=False,
    ):
        """Initialize the RAJson class instance"""
        self.batch_id = batch_id
        self.input_path = input_path
        self.info_holder_data = info_holder_data
        self.valid_page_types = valid_page_types
        self.input_folder_path = os.path.join(self.input_path, self.batch_id)
        self.excel_read_only = excel_read_only
        self.read_checkbox = read_checkbox
        self.ra_json_error = None

    def process_element(self, element):
        """
        This function reads attributes of given XML Tree element and return details in form of dictionary.
        Children property is generated recursively.

        Args:
            element: The XML element to process.

        Returns:
            dict: Contain the attributes and structure of the XML element.

        Process Details:
            - Extracts attributes from the element and maps its type using 'self.NODES'.
            - Handles child elements recursively, filtering based on 'self.NODE_KEYS'.
            - Adds an empty 'id' field to the element's attributes.
            - Includes a 'children' field in the dictionary for non-leaf nodes.

        Notes:
            - Leaf nodes of type 'word' do not have children.
            - The 'type' field is determined using 'self.NODES' mapping or defaults to the element's tag.
            - The 'children' field is populated only if child elements are present.
        """
        element_type = element.tag
        attrs = element.attrib
        attrs["type"] = self.NODES.get(element_type, element_type)

        if attrs["type"] == "word":
            children = []
        else:
            child_elements = [i for i in list(element) if i.tag in self.NODE_KEYS]
            children = [self.process_element(e) for e in child_elements]

        attrs["id"] = ""
        if children:
            attrs["children"] = children
        return attrs

    def process_PAGE(self, xml_file_path):
        """
        This function accepts path of XML file on disk, convert XML file into tree object,
        reads attributes root (PAGE) element and return details in form of dictionary.
        Children property is generated recursively.

        Args:
            xml_file_path (str): The file path of the XML file to be processed.

        Returns:
            dict: Contains the attributes and structure of the PAGE element.

        Process Details:
            - Parses the XML file and retrieves the root PAGE element.
            - Extracts the attributes of the PAGE element and assigns "page" as its type.
            - Extracts style information from "Style" nodes and stores it in the 'styles' key.
            - Filters and processes child elements recursively using 'self.process_element'.

        Notes:
            - Only child elements matching 'self.NODE_KEYS' are processed.
            - Style nodes are extracted separately and do not appear in the 'children' list.
        """
        tree = ET.parse(xml_file_path)
        PAGE = tree.getroot()
        attrs = PAGE.attrib
        attrs["type"] = "page"

        childrens = list(PAGE)

        # Add style info to styles key in page
        style_nodes = [i for i in childrens if i.tag == "Style"]
        attrs["styles"] = [i.attrib for i in style_nodes]

        # Filter only required elements
        childrens = [i for i in childrens if i.tag in self.NODE_KEYS]
        attrs["children"] = [self.process_element(e) for e in childrens]

        return attrs

    def pre_process_PAGE(self, P):
        """
        This function accepts Page element and generate XML file path based on Page ID.
        Then each page is processed using process_PAGE function.

        Args:
            P: The PAGE element from the XML tree.

        Returns:
            dict: Contains the combined attributes of the PAGE element.

        Process Details:
            - Extracts the 'id' attribute of the PAGE element.
            - Reads export attributes from child "V" elements of the PAGE.
            - Constructs the XML file path based on the PAGE ID.
            - Processes the XML file using 'self.process_PAGE' to retrieve its hierarchical structure.
            - Combines export attributes with the processed XML attributes.

        Notes:
            - The 'IMAGEFILE' attribute is converted to lowercase for consistency.
            - The XML file path is assumed to follow the format '{page_id}_layout.xml'.
        """
        page_id = P.get("id")
        export_attrs = {}
        for V in P.findall("./V"):
            key = V.get("n")

            # always rename imagefile to lowercase
            if key == "IMAGEFILE":
                export_attrs[key] = V.text.lower()
            else:
                export_attrs[key] = V.text

        xml_file_path = os.path.join(self.input_folder_path, f"{page_id}_layout.xml")

        page_attrs = self.process_PAGE(xml_file_path)
        page_attrs.update(export_attrs)
        return page_attrs

    def _process_shapes_excel_xml(self, excel_file_path):
        shape_dict = {}

        with zipfile.ZipFile(excel_file_path, "r") as zip_ref:
            drawing_files = [
                f for f in zip_ref.namelist() if f.startswith("xl/drawings/")
            ]

            for drawing_file in drawing_files:
                with zip_ref.open(drawing_file) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()

                    namespaces = {
                        "v": "urn:schemas-microsoft-com:vml",
                        "o": "urn:schemas-microsoft-com:office:office",
                        "x": "urn:schemas-microsoft-com:office:excel",
                    }

                    vshape_elements = root.findall(".//v:shape", namespaces)

                    for vshape in vshape_elements:
                        shape_id = vshape.get("id")

                        checked_element = vshape.find(".//x:Checked", namespaces)
                        checked = (
                            checked_element.text == "1"
                            if checked_element is not None
                            else False
                        )

                        textbox = vshape.find(".//v:textbox", namespaces)
                        if textbox is not None:
                            div = textbox.find(".//div")
                            if div is not None:
                                text = "".join(div.itertext()).strip()
                            else:
                                text = ""
                        else:
                            text = ""
                        if text != "":
                            shape_dict[shape_id] = {"checked": checked, "text": text}
        return shape_dict

    def _update_shipment_type(
        self,
        shipment_type_dict: dict,
        value: str,
        row_idx: int,
        column_letter: str,
        original_value: str,
    ) -> dict:
        shipment_type_dict["value"] = value
        shipment_type_dict["original_value"] = original_value
        shipment_type_dict["column_letter"] = column_letter
        shipment_type_dict["row_idx"] = row_idx
        return shipment_type_dict

    def _filter_non_fcl_values_if_exists(self, all_checked_values):
        fcl_is_only_checked = all([i["value"] == "FCL" for i in all_checked_values])
        if fcl_is_only_checked and len(all_checked_values) > 1:
            # If fcl is only checked one is default one is from user just return the first one
            return all_checked_values[:1]
        if fcl_is_only_checked and len(all_checked_values) == 1:
            # If nothing is checked fcl will be default checked and its length will be one just return 0
            return []
        if not fcl_is_only_checked and len(all_checked_values) > 1:
            """
            # This case is handeled because FCL was checked from nowhere from excel
            # Case 1: `FCL 整柜` -> FCL with space and chinese_character was showing checkbox with default value as it is.
            # Case 2: `FCL整柜` -> FCL without chinese character was auto detected as true and taken out in `all_checked_values` so to avoid the case if FCL is present but if any other values are present then extract either BCN or LCL else extract FCL
            CASE 1:`FCL 整柜` -> FCL with space and chinese_character was showing checkbox with default value and auto true as it is.
            [
                {
                    "id": "1. DGF CN SLI Sample.L:25",
                    "value": "FCL",
                    "original_value": "FCL整柜",
                    "column_letter": "L",
                    "row_idx": 25
                },
                {
                    "id": "1. DGF CN SLI Sample.L:25",
                    "value": "FCL",
                    "original_value": "FCL 整柜",
                    "column_letter": "L",
                    "row_idx": 25
                }
            ]
            For this case FCL is extracted no matter other keywords
            Case 2: `FCL整柜` -> FCL without chinese character was auto detected as true
            [
                {
                    "id": "1. DGF CN SLI Sample.L:25",
                    "value": "FCL",
                    "original_value": "FCL整柜",
                    "column_letter": "L",
                    "row_idx": 25
                },
                {
                    "id": "1. DGF CN SLI Sample.L:26",
                    "value": "LCL",
                    "original_value": "LCL 散货",
                    "column_letter": "L",
                    "row_idx": 26
                }
                ]
            For this case FCL is extracted since LCL is checked we will get LCL only
            """
            count_fcl = sum(1 for i in all_checked_values if i["value"] == "FCL")
            if count_fcl == 1:
                # Remove all occurrences of "FCL" if it is default checked from excel
                all_checked_values = [
                    i for i in all_checked_values if i["value"] != "FCL"
                ]
            elif count_fcl > 1:
                # Remove only one occurrence of "FCL" the one will be default checked and other will be checked manually by user.
                removed = False
                all_checked_values = [
                    i
                    for i in all_checked_values
                    if not (i["value"] == "FCL" and not removed and (removed := True))
                ]
        return all_checked_values

    def _get_all_checked_values(self, excel_file_path, sheet_name):
        shape_dict = self._process_shapes_excel_xml(excel_file_path)
        all_checked_values = []
        with zipfile.ZipFile(excel_file_path, "r") as z:
            drawing_files = [
                f for f in z.namelist() if f.startswith("xl/drawings/drawing")
            ]

            for drawing_file in drawing_files:
                with z.open(drawing_file) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()

                    ns = {
                        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                        "ns0": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                        "ns1": "http://schemas.openxmlformats.org/drawingml/2006/main",
                        "ns2": "http://schemas.microsoft.com/office/drawing/2010/main",
                        "ns3": "http://schemas.openxmlformats.org/markup-compatibility/2006",
                    }
                    for idx, checkbox in enumerate(
                        root.findall(".//xdr:twoCellAnchor", ns)
                    ):
                        text_element = checkbox.find(".//a:t", ns)
                        if text_element is not None:
                            checkbox_label = text_element.text.strip()
                            compat_ext = checkbox.find(".//ns2:compatExt", ns)
                            if compat_ext is not None:
                                spid = compat_ext.get("spid")
                                if spid in shape_dict:
                                    if shape_dict[spid]:
                                        checked = shape_dict[spid].get("checked", False)
                                        if checked:
                                            from_cell = checkbox.find(".//xdr:from", ns)
                                            if from_cell is not None:
                                                col_idx = (
                                                    int(
                                                        from_cell.find(
                                                            "xdr:col", ns
                                                        ).text
                                                    )
                                                    + 1
                                                )

                                                row_idx = (
                                                    int(
                                                        from_cell.find(
                                                            "xdr:row", ns
                                                        ).text
                                                    )
                                                    + 1
                                                )

                                                column_letter = get_column_letter(
                                                    col_idx
                                                )
                                                shipment_type_dict = {
                                                    "id": f"{sheet_name}.{column_letter}:{row_idx}",
                                                    "value": checkbox_label,
                                                }
                                                update_shipment_type_kwargs = {
                                                    "column_letter": column_letter,
                                                    "row_idx": row_idx,
                                                    "shipment_type_dict": shipment_type_dict,
                                                    "original_value": checkbox_label,
                                                }
                                                shipment_types = ["BCN", "LCL", "FCL"]
                                                for shipment in shipment_types:
                                                    if (
                                                        shipment in checkbox_label
                                                        and checked
                                                    ):
                                                        all_checked_values.append(
                                                            self._update_shipment_type(
                                                                **update_shipment_type_kwargs,
                                                                value=shipment,
                                                            )
                                                        )

        all_checked_values = self._filter_non_fcl_values_if_exists(all_checked_values)

        return all_checked_values

    def _is_exception_excel_checkbox_profile(self, batch_data):
        """Check if current profile is a exception excel checkbox profiles"""
        profile_name = batch_data["DefinitionID"]

        if not profile_name or not ApplicationSettings.objects.exists():
            return False

        application_settings = ApplicationSettings.objects.first().data
        exception_excel_checkbox_profiles = application_settings["profileSettings"].get(
            "exception_excel_checkbox_profiles", []
        )
        exception_excel_checkbox_profiles = [
            i["name"] for i in exception_excel_checkbox_profiles
        ]

        if profile_name in exception_excel_checkbox_profiles:
            return True

        return False

    def process_excel_sheet(self, sheet_obj, excel_file_path, batch_data):
        """
        Extract information for multiple sheets from give excel workbook.
        Extracts its metadata and cell content, and organizes the information into a structured dictionary format.
        Non-empty cells are stored along with their coordinates and values.

        Args:
            sheet_obj: The worksheet object to be processed.

        Returns:
            dict: Contains the worksheet's metadata and cell data, where each column is represented as a key,
                  and each row is a subkey.

        Process Details:
            - Extracts the worksheet title as metadata.
            - Iterates through cell values in the worksheet, skipping empty cells.
            - Constructs a unique ID for each cell based on its sheet name, column letter, and row index.
            - Serializes cell data to handle non-serializable types.

        Dependencies:
            - get_column_letter from openpyxl.utils

        Notes:
            - Empty cells are ignored during processing.
            - Non-serializable data types are converted to JSON-compatible formats.
        """
        attrs = {}
        sheet_name = sheet_obj.title
        attrs["id"] = sheet_name
        attrs["type"] = "worksheet"
        attrs["title"] = sheet_name

        cell_data = {}
        # Get only the values without cell objects
        for row_idx, row in enumerate(sheet_obj.values, start=1):
            for col_idx, value in enumerate(row, start=1):
                if value is not None:  # Skip empty cells
                    column_letter = get_column_letter(col_idx)
                    cell_item = {
                        "id": f"{sheet_name}.{column_letter}:{row_idx}",
                        "value": value,
                    }
                    try:
                        cell_data[column_letter][row_idx] = cell_item
                    except KeyError:
                        # First time we see this column
                        cell_data[column_letter] = {}
                        cell_data[column_letter][row_idx] = cell_item

        if self.read_checkbox and self._is_exception_excel_checkbox_profile(batch_data):
            all_checked_values = self._get_all_checked_values(
                excel_file_path, sheet_name
            )

            if len(all_checked_values) != 1:
                raise ValueError("Checkbox should have exactly one value checked")

            first_checked_value = all_checked_values[0]
            if first_checked_value:
                cell_data.setdefault(first_checked_value["column_letter"], {})[
                    first_checked_value["row_idx"]
                ] = {
                    "id": first_checked_value["id"],
                    "value": first_checked_value["value"],
                }

        # Dump and load json to handle non serializable date (i.e. datetime)
        cell_data = json.loads(json.dumps(cell_data, default=json_default))
        attrs["cell_data"] = cell_data

        return attrs

    def process_D(self, D, batch_data):
        """
        This function accepts Document element, and process each page within the
        document using pre_process_PAGE function.

        Args:
            D : The document element from the XML tree.
            batch_data (dict): A dictionary containing batch-level metadata.

        Returns:
            dict: containing the document ID, type, extension, and a list of processed pages
                  or sheets along with batch metadata.

        Raises:
            ValueError: If the document extension is '.xls'.

        Process Details:
            - Extracts the document's ID and type.
            - Determines the document's file extension, defaulting to '.pdf' if unspecified.
            - Handles unsupported extensions like '.xls' by raising exceptions.
            - Filters valid pages based on their 'TYPE' attributes for '.pdf', '.doc' and '.docx' files.
            - Processes valid pages using the 'pre_process_PAGE' function.
            - Processes '.xlsx' files by extracting sheet data using the 'process_excel_sheet' function.

        Dependencies:
            - load_workbook from openpyxl

        Notes:
            - '.xls' files are explicitly unsupported and raise a 'ValueError'.
            - Any extension other than '.doc', '.docx', '.xls', '.xlsx', and '.pdf' is defaulted to '.pdf'.
            - If no extension is specified, '.pdf' is assumed.
        """
        doc_id = D.get("id")
        attrs = {"type": "document", "id": doc_id}

        attrs.update(batch_data)

        # Capture extension value
        for V in D.findall("./V"):
            key = V.get("n")
            value = V.text
            if key.lower() == "ext":
                # Empty variables will be considered as .pdf
                if value is None:
                    attrs["ext"] = ".pdf"

                elif value in ["xls", ".xls"]:
                    attrs["ext"] = ".xls"
                    raise ValueError("Unsupported document extension .xls")

                elif value in ["xlsx", ".xlsx"]:
                    attrs["ext"] = ".xlsx"

                elif value in ["doc", ".doc", "docx", ".docx"]:
                    attrs["ext"] = ".docx"

                #  Anything other than doc, docx, xls and xlsx will be considered as .pdf
                else:
                    attrs["ext"] = ".pdf"

                break

        # If ext variable is not available, consider it as .pdf
        document_extention = attrs.get("ext", ".pdf")

        if document_extention in [".pdf", ".docx"]:
            pages = D.findall("P")
            valid_pages = []

            for P in pages:
                for V in P.findall("./V"):
                    key = V.get("n")
                    value = V.text
                    if key == "TYPE":
                        if value in self.valid_page_types:
                            valid_pages.append(P)
                            break

            attrs["children"] = [self.pre_process_PAGE(P) for P in valid_pages]

            return attrs

        elif document_extention == ".xlsx":
            # Get excel file name from first P node
            P = D.find("P")
            for V in P.findall("./V"):
                key = V.get("n")
                value = V.text
                if key == "IMAGEFILE":
                    attrs["sourceFileName"] = os.path.splitext(value)[0].lower()
                    break

            source_file_name = attrs["sourceFileName"]
            excel_file_name = f"{source_file_name}.xlsx"
            # if monitor passes then the files are good to parse.
            excel_file_path = os.path.join(self.input_folder_path, excel_file_name)
            wb = load_workbook(
                excel_file_path, data_only=True, read_only=self.excel_read_only
            )
            sheets = wb.worksheets
            try:
                # Add sheetnames if it matches with profile
                profile_name = batch_data["DefinitionID"]
                profile = Profile.objects.get(name=profile_name)
                if profile.multi_shipment:
                    attrs["sheetnames"] = wb.sheetnames
            except:
                pass
            attrs["children"] = [
                self.process_excel_sheet(i, excel_file_path, batch_data) for i in sheets
            ]
            return attrs

    def get_file_name(self, excel_file_name):
        """
        Retrieves the file name associated with the current batch ID from the database.

        Args:
            excel_file_name (str): The fallback file name to return if no matching
            document is found in the database.

        Returns:
            str: The file name retrieved from the database or the provided 'excel_file_name'.

        Process Details:
            - Queries 'EmailParsedDocument' for documents matching the 'batch_id'.
            - Queries 'TrainParsedDocument' if no match is found in 'EmailParsedDocument'.
            - Returns the 'name' of the first matching document or the provided fallback file name.

        Dependencies:
            - EmailParsedDocument, TrainParsedDocument from core.models

        Notes:
            - Relies on class-level attribute 'batch_id' for filtering.
            - Assumes that both 'EmailParsedDocument' and 'TrainParsedDocument' models have a 'name' field.
        """
        email_document = EmailParsedDocument.objects.filter(batch_id=self.batch_id)
        train_document = TrainParsedDocument.objects.filter(batch_id=self.batch_id)

        if email_document.exists():
            return email_document.first().name
        elif train_document.exists():
            return train_document.first().name
        else:
            return excel_file_name

    def process_export_file(self):
        """
        This function finds Export xml file from input_folder_path, convert the xml
        file to tree elements and process each document within it using process_D function.

        Args:
            Uses class-level attributes like 'info_holder_data' and 'input_folder_path'.

        Returns:
            dict (attrs) : processed export file's attributes, metadata, and structured document nodes.

        Process Details:
            - Parses the XML file located in the 'input_folder_path'.
            - Extracts key metadata such as profile, vendor, document type, language, project,
              and name matching text from the XML's "V" element.
            - Processes document elements ("D" nodes) using the 'process_D' function to
              create a list of structured document nodes.
            - Determines the batch type (file extension) from the first document node.

        Notes:
            - Relies on 'info_holder_data' for variable name mappings (profile, vendor, etc.).
            - Uses 'ET.parse()' to read and parse the XML structure.
            - Iterates over "V" elements for metadata and "D" elements for document processing.
        """
        layout_file_name = self.info_holder_data["name"]
        export_file = os.path.join(self.input_folder_path, layout_file_name)

        tree = ET.parse(export_file)
        B = tree.getroot()

        attrs = {"id": B.get("id")}

        #  Get profile (definition_id), vendor, document type from xml
        profile_var_name = self.info_holder_data["profileVarName"]
        vendor_var_name = self.info_holder_data["vendorVarName"]
        doc_type_var_name = self.info_holder_data["docTypeVarName"]
        language_var_name = self.info_holder_data["languageVarName"]
        project_var_name = self.info_holder_data["projectVarName"]
        name_matching_text_var_name = self.info_holder_data["nameMatchingTextVarName"]

        batch_data = {}

        for V in B.findall("V"):
            key = V.get("n")
            value = V.text
            if key == profile_var_name:
                attrs["DefinitionID"] = value
                batch_data["DefinitionID"] = value

            elif key == vendor_var_name:
                attrs["Vendor"] = value
                batch_data["Vendor"] = value

            elif key == doc_type_var_name:
                attrs["DocumentType"] = value
                batch_data["DocType"] = value

            elif key == language_var_name:
                attrs["Language"] = value
                batch_data["Language"] = value

            elif key == project_var_name:
                attrs["Project"] = value
                batch_data["Project"] = value

            elif key == name_matching_text_var_name:
                attrs["NameMatchingText"] = value
                batch_data["NameMatchingText"] = value

            else:
                attrs[key] = value

        # for V in tree.findall("./V"):
        #     key = V.get("n")
        #     attrs[key] = V.text

        attrs["nodes"] = [self.process_D(D, batch_data) for D in B.iter("D")]

        # Get extenstion from first document
        attrs["batch_type"] = attrs["nodes"][0].get("ext", ".pdf")

        return attrs

    def assign_id(self, element):
        """
        This function accepts dictionary object. Based on ID of the provided dictionary object,
        it recursively provides ID to all its nested (child) elements.

        Args:
            element (dict): A dictionary object contain an 'id' key and optionally a 'children' key for nested elements.

        Returns:
            The dictionary is updated with assigned IDs.

        Process Details:
            - The function starts with the 'id' of the given element ('base_id').
            - If the element has a 'children' key, it iterates through each child.
            - Each child is assigned a unique ID in the format '<base_id>.<index>', where the index is
              zero-padded to three digits (e.g., '1.001', '1.002').
            - The function calls itself recursively for further nested children.

        Notes:
            - The IDs follow a hierarchical numbering format to reflect the parent-child relationship.
            - This function modifies the original dictionary object directly.
        """
        base_id = element["id"]
        children = element.get("children")
        if children:
            for index, child in enumerate(children):
                child_id = f"{base_id}.{str(index + 1).zfill(3)}"
                child["id"] = child_id
                self.assign_id(child)

    def create_DB_json(self, result_json):
        """
        This function is responsible for creating Doc Builder JSON from Original JSON. [not USING right now]
        The Doc Builder JSON only contains elemets of Line and Words Type.

        Args:
            result_json (dict): Original JSON structure containing document data.

        Returns:
            The 'result_json' dictionary is updated with additional Doc Builder nodes.

        Process Details:
            - Each document in the original JSON is duplicated and prefixed with "DB" in its IDs.
            - Only elements of type 'line' and their child elements of type 'words' are processed.
            - Words within a 'line' element are concatenated into a single string,
              and a new child of type 'words' is added to the line.
            - A helper function 'process_item' recursively traverses elements and processes lines.

        Notes:
            - The original JSON structure is not altered; a new set of Doc Builder nodes is added to it.
            - 'process_item(element)': Recursively processes child elements, identifying 'line' types
              and assigning them to the Doc Builder structure.
            - This function uses a deep copy of the original JSON to preserve the hierarchy
              while transforming it for Doc Builder.
        """
        DB_DOC_lines = []
        page_id = ""

        def process_item(element):
            """
            Update ID, add page ID for "line" types element
            """
            element["id"] = "DB" + element["id"]
            if element["type"] == "line":
                element["pageId"] = page_id
                DB_DOC_lines.append(element)
            else:
                if "children" in element.keys():
                    for item in element["children"]:
                        process_item(item)

        DB_json = deepcopy(result_json)
        DB_documents = DB_json["nodes"]
        for document in DB_documents:
            document["id"] = "DB" + document["id"]

            pages = document["children"]
            document["children"] = []
            for page in pages:
                page_id = page["id"]

                DB_DOC_lines = []
                process_item(page)

                for line in DB_DOC_lines:
                    words = line["children"]
                    words_list = [word["v"] for word in words]
                    words_string = " ".join(words_list)
                    line["children"] = [
                        {
                            "id": line["id"] + "_words",
                            "type": "words",
                            "pos": line["pos"],
                            "v": words_string,
                            "pageId": page_id,
                        }
                    ]

                document["children"] += DB_DOC_lines
                document["type"] = "docBuilder"

        result_json["nodes"] = result_json["nodes"] + DB_documents

    def process_folder(self):
        """
        This function is wrapper to all processes.
        First it generates dictionary object for all XML files in given input_folder_path.
        Then it calls assign_id function to recursively provide IDs to all child elements of PAGE Elements.
        Then dumps JSON to the disk.

        Args:
            Operates on class attributes.

        Returns:
            dict: A dictionary object representing the processed file structure, with assigned IDs.

        Process Details:
            - The 'process_export_file' function is called to generate a JSON-like structure containing file nodes.
            - For files with '.pdf' or '.docx' extensions, each page is processed using 'assign_id' to ensure unique identification.

        Notes:
            - The function assume that 'result_json' follow a specific schema where:
                - 'nodes' is a list of documents.
                - Each document has an 'ext' key for file extension and a 'children' key for pages.
            - Only files with '.pdf' or '.docx' extensions are considered for ID assignment.
        """
        result_json = self.process_export_file()

        for document in result_json["nodes"]:
            extension = document.get("ext", ".pdf")
            if extension in [".pdf", ".docx"]:
                for page in document["children"]:
                    self.assign_id(page)

        # Skip generating DB JSONs
        # self.create_DB_json(result_json)

        return result_json

    def process(self):
        """Execute the main processing workflow for the class"""
        result_json = self.process_folder()
        return result_json
